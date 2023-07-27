import os
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
import shutil
from shutil import copy2
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from pdf2image import convert_from_path
import pytesseract
from extract_msg import Message
from PyPDF2 import PdfReader, PdfWriter, PdfFileWriter

from LinkCrudeResourcesLLC import extract_data_link_crude
from CirtronCommoditiesLLC import extract_data_citron_commodities
from ModernCommoditiesINC import extract_data_modern_commodities
from OneExchangeCorp import extract_data_one_exchange
from CalRockBrokersINC import extract_data_calrock_brokers
from SyntexEnergyLLC import extract_data_syntex_energy
from MarexSpectron import extract_data_marex_spectron

# Path for tesseract
poppler_path = r"C:\Program Files\poppler-0.68.0_x86\poppler-0.68.0\bin"  # replace with your path
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
os.environ["PATH"] += os.pathsep + poppler_path

# Define the names & locations of folders and format / data files
BASE_DIR = os.getcwd()
PDF_DIR = os.path.join(BASE_DIR, "pdfs")
DATA_DIR = os.path.join(BASE_DIR, "data")
RESULT_DIR = os.path.join(BASE_DIR, "results")
FORMAT_FILE = os.path.join(BASE_DIR, "format.xlsx")
physical_data_locations = os.path.join(BASE_DIR, "physical_data_locations.xlsx")
recognization = {}

# Copy every excel from one folder to another
def copy_files(source_dir, target_dir):
    if not os.path.exists(target_dir):
        os.makedirs(target_dir)
    
    files = [f for f in os.listdir(source_dir) if os.path.isfile(os.path.join(source_dir, f))]
    
    for file in files:
        if file.endswith('.xlsx'):
            copy2(os.path.join(source_dir, file), target_dir)

# Copy the format.xlsx file to each excel file
def copy_format_to_each_file(target_dir, format_file):
    files = [f for f in os.listdir(target_dir) if os.path.isfile(os.path.join(target_dir, f)) and f.endswith('.xlsx')]

    format_df = pd.read_excel(format_file)
    
    for file in files:
        book = load_workbook(os.path.join(target_dir, file))

        if 'Sheet2' in book.sheetnames:
            sheet = book['Sheet2']  # If 'Sheet2' already exists, select it
        else:
            sheet = book.create_sheet('Sheet2')  # Otherwise, create 'Sheet2'

        for row_index, row in format_df.iterrows():
            for column_index, cell_value in enumerate(row):
                # +1 is added to row and column indices because Openpyxl's index starts from 1, not 0
                sheet.cell(row=row_index+1, column=column_index+1, value=cell_value)
        
        book.save(os.path.join(target_dir, file))

# Helper function
def load_files(directory, extension=None):
    if extension:
        return [f for f in os.listdir(directory) if os.path.isfile(os.path.join(directory, f)) and f.endswith(extension)]
    else:
        return [f for f in os.listdir(directory) if os.path.isfile(os.path.join(directory, f))]


broker_to_function_map = {
    "Broker Link Crude": extract_data_link_crude,
    "Broker Citron Commodities": extract_data_citron_commodities,
    "Broker CalRock Brokers": extract_data_calrock_brokers,
    "Broker Modern Commodities": extract_data_modern_commodities,
    "Broker One Exchange": extract_data_one_exchange,
    "Broker Syntex Energy": extract_data_syntex_energy,
    "Broker Marex Spectron": extract_data_marex_spectron,
    # Add more broker and values here...
}

def identify_broker(sheet):
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str):
                if 'LINK CRUDE RESOURCES, LLC' in cell:
                    print("found Broker Link Crude")
                    return 'Broker Link Crude'
                if 'None None' in cell:
                    print("found Broker Citron Commodities")
                    return 'Broker Citron Commodities'
                if 'CalRock Brokers Inc.' in cell:
                    print("found Broker CalRock Brokers")
                    return 'Broker CalRock Brokers'
                if 'Click & Trade' in cell:
                    print("found Broker Modern Commodities")
                    return 'Broker Modern Commodities'
                if 'ONE EXCHANGE' in cell:
                    print("found Broker One Exchange")
                    return 'Broker One Exchange'
                if 'One Exchange' in cell:
                    print("found Broker One Exchange")
                    return 'Broker One Exchange'
                if 'SYNTEXENERGY' in cell:
                    return 'Broker Syntex Energy'
                if 'Syntex Energy' in cell:
                    return 'Broker Syntex Energy'
                if 'Marex' in cell:
                    return 'Broker Marex Spectron'

    # If no broker found, return None
    print("Broker not found")
    return None

def extract_data(sheet, file_name):
    brokerCompany = identify_broker(sheet)
    
    if brokerCompany is None:
        print(f"Could not identify broker in file: {file_name}")
        return None

    # Use the appropriate extraction function for the identified broker
    extraction_function = broker_to_function_map.get(brokerCompany)
    if extraction_function:
        return extraction_function(sheet)
    else:
        print(f"Could not find extraction function for broker: {brokerCompany}")


# Printing the data collected to the corresponding excel files, including constant data
def update_sheet(sheet, data, filename):
    transaction_date, transaction_type, seller, buyer, pipeline, location, trader, quantityA, quantityB, quantityC, broker, brokerDocID, \
    pricingDetail, pricingType, premium, paymentTerm, creditTerm, delivery_date_start, delivery_date_end, id_, team, currency, deliveryTerm = data or ""

    sheet['B1'] = transaction_date or ""
    sheet['B3'] = transaction_type or ""
    #print(f"buyer is", buyer)
    #print(f"seller is", seller)
    if buyer and 'PETROCHINA' in buyer:
        #print(f"we are buying, trader is", trader)
        sheet['B2'] = 'Buy'
        sheet['B4'] = buyer
        sheet['B5'] = seller
        sheet['B14'] = trader or ""
    else:
        #print("we are selling")
        sheet['B2'] = 'Sell'
        sheet['B4'] = seller
        sheet['B5'] = buyer
        sheet['B14'] = trader or ""
        quantityA = -1 * (quantityA or 0)
    try:
        sheet['B8'] = (delivery_date_start.strftime('%m/%d/%Y').lstrip("0").replace("/0", "/") if delivery_date_start else "")
        sheet['B9'] = (delivery_date_end.strftime('%m/%d/%Y').lstrip("0").replace("/0", "/") if delivery_date_end else "")
    except AttributeError:
        sheet['B8'] = delivery_date_start
        sheet['B9'] = delivery_date_end

    sheet['B10'] = quantityA or ""
    sheet['B11'] = quantityB or ""
    sheet['B12'] = quantityC or ""
    sheet['B13'] = deliveryTerm
    sheet['B15'] = team
    sheet['B17'] = 'Pipeline'
    sheet['B18'] = location or ""  
    sheet['B19'] = pipeline or ""

    try:
        sheet['B22'] = (delivery_date_start.strftime('%Y-%m-%d') if delivery_date_start else "")
    except AttributeError:
        print(delivery_date_start)
        sheet['B22'] = ''

    sheet['B23'] = pricingType or ""
    sheet['B24'] = pricingDetail or ""
    sheet['B25'] = premium or ""
    sheet['B26'] = paymentTerm or ""
    sheet['B27'] = creditTerm or ""
    sheet['B28'] = 'N/A'
    sheet['B29'] = 'No Inspection Fee Associated'
    sheet['B30'] = '0'
    sheet['B32'] = currency
    sheet['B33'] = broker or ""
    sheet['B34'] = brokerDocID or ""
    sheet['B35'] = id_ or ""
    
    if not sheet['B14'].value:
        recognization[filename] = False
    if sheet['B35'].value == 'no corresponding pipeline implis no correct id':
        recognization[filename] = False


def copy_format_to_sheet(format_file_name, sheet):
    format_book = load_workbook(format_file_name)
    format_sheet = format_book.active

    for i, column in enumerate(format_sheet.columns, start=1):
        column_letter = get_column_letter(i)
        if column_letter in format_sheet.column_dimensions:
            format_width = format_sheet.column_dimensions[column_letter].width
            sheet.column_dimensions[column_letter].width = format_width

def cleanup_data(target_dir):
    files = load_files(target_dir)

    for file in files:
        book = load_workbook(os.path.join(target_dir, file))
        for sheet_name in book.sheetnames:
            sheet1 = book[sheet_name]
            
            # If the sheet has only two rows of data
            if sheet1.max_row == 2:
                transformed_data = []
                for row in sheet1.iter_rows(min_row=1, max_row=2, values_only=True):
                    transformed_data.append(row)
                sheet1.delete_rows(1, 2)
                for idx, (item_A, item_B) in enumerate(zip(transformed_data[0], transformed_data[1]), start=1):
                    sheet1.cell(row=idx, column=1, value=f"{item_A} : {item_B}")
                book.save(os.path.join(target_dir, file))

            # Existing functionality
            elif sheet1.max_column == 2:
                for row in sheet1.iter_rows(min_row=2, max_col=2, max_row=sheet1.max_row):
                    cell_A = row[0]
                    cell_B = row[1]
                    cell_A.value = f"{cell_A.value} {cell_B.value}"
                    cell_B.value = None
                book.save(os.path.join(target_dir, file))

        data_dict = extract_data(sheet1, file)

        if data_dict is None:
            print(f"Skipping file {file} due to broker identification issue.")
            continue  # Skip to the next file

        if 'Sheet2' in book.sheetnames:
            sheet2 = book['Sheet2']
        else:
            sheet2 = book.create_sheet('Sheet2')

        copy_format_to_sheet('format.xlsx', sheet2)
        update_sheet(sheet2, data_dict, file)
        book.save(os.path.join(target_dir, file))

# If the pdf has two pages, split them (its usally a buy-trade so they're put together, but they are recorded seperately)
def split_pdf_pages(directory):
    for file in os.listdir(directory):
        if file.endswith('.pdf'):
            pdf_path = os.path.join(directory, file)
            #print(f"Processing: {file}")
            pdf = PdfReader(pdf_path)

            if len(pdf.pages) >= 2:
                for i, page in enumerate(pdf.pages, start=1):
                    writer = PdfWriter()
                    writer.add_page(page)
                    output_file = os.path.join(directory, f"{file.replace('.pdf', '')}_{i}.pdf")
                    writer.write(output_file)
                    print(f"Created: {output_file}")
                os.remove(pdf_path)
                print(f"Deleted original file: {file}")


# PDF Plumber that changes broker pdf to excel
def extract_text_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = []
        for page in pdf.pages:
            page_text = page.extract_text()
            for line in page_text.split('\n'):
                text.append([line])  # wrap each line in a list to create a 2D list

    if not text or all(line == [''] for line in text):
        print('PdfPlumber Failed')
        return None

    return text


# Tesseract in case PDF Plumber fails
def pdf_to_excel(pdf_file, output_file):
    images = convert_from_path(pdf_file)
    all_text = []
    for i, image in enumerate(images):
        #print(f"Processing page {i+1}...")
        
        image = image.convert('L')  # Improve image quality, convert image to grayscale
        text = pytesseract.image_to_string(image, config='--oem 1 --psm 3')         # Extract text using OCR
        #print(text)  # Print out the text
        text = text.split('\n')
        all_text.extend(text)
    
    df = pd.DataFrame(all_text, columns=["Text"])       # Create DataFrame
    #print(df.head())                                    # Check DataFrame
    df.to_excel(output_file, index=False)               # Export to Excel

# Helper function
def write_to_excel(data, output_file):
    df = pd.DataFrame(data, columns=['Text'])
    df.to_excel(output_file, index=False)

# Function that deletes the data & results folder each time before the new info is recorded
def remove_all_files_from_directory(directory):
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s' % (file_path, e))

# Fuction to process pdf files
def process_pdf_files(directory):
    for file in os.listdir(directory):
        if file.endswith('.pdf'):
            pdf_path = os.path.join(directory, file)
            output_file = os.path.join(DATA_DIR, file.replace('.pdf', '.xlsx'))
            text = extract_text_from_pdf(pdf_path)                  # Try to extract text using pdfplumber
            if text is None: 
                print('PdfPlumber Failed, trying Tesseract')
                pdf_to_excel(pdf_path, output_file)    # Try to extract text using tesseract
            else: write_to_excel(text, output_file)                 # Write the extracted text to Excel

# Function to process excel files
def process_xlsx_files(directory):
    for file in os.listdir(directory):
        if file.endswith('.xlsx'):
            original_wb = load_workbook(os.path.join(directory, file))
            for i, sheet in enumerate(original_wb.sheetnames, start=1):
                wb = Workbook()
                new_sheet = wb.active
                for row in original_wb[sheet].iter_rows():
                    new_sheet.append((cell.value for cell in row))
                output_file = os.path.join(DATA_DIR, f"{file.replace('.xlsx', '')}_{i}.xlsx")
                wb.save(output_file)

# Helper function to process msg files
def convert_msg_to_excel(msg_file_path):
    msg = Message(msg_file_path)
    msg_content = msg.body

    # Split the message into lines and remove excess whitespace
    msg_lines = msg_content.split('\n')
    cleaned_msg_lines = [' '.join(line.split()) for line in msg_lines]

    return cleaned_msg_lines

# Function to process msg files
def process_msg_files(directory):
    for file in os.listdir(directory):
        if file.endswith('.msg'):
            msg_file_path = os.path.join(directory, file)
            msg_lines = convert_msg_to_excel(msg_file_path)
            
            # Create a list to store message lines and remove excess spaces
            message_lines = [' '.join(line.split()) for line in msg_lines]

            # Create a DataFrame from the list
            df = pd.DataFrame({'Message Line': message_lines})

            # Write the DataFrame to an Excel file
            output_file = os.path.join(DATA_DIR, file.replace('.msg', '.xlsx'))
            df.to_excel(output_file, index=False)


# Main(), where all the functions are called and handles the command line
# Simply run 'python dataCleanup.py' on terminal after having the dataCleanup folder open
def main():
    for directory in [DATA_DIR, PDF_DIR, RESULT_DIR]:
        if not os.path.exists(directory):
            os.makedirs(directory)
    remove_all_files_from_directory(DATA_DIR)
    remove_all_files_from_directory(RESULT_DIR)
    
    split_pdf_pages(PDF_DIR)
    process_pdf_files(PDF_DIR)
    process_xlsx_files(PDF_DIR)
    process_msg_files(PDF_DIR)

    cleanup_data(DATA_DIR)
    copy_files(DATA_DIR, RESULT_DIR)
    copy_format_to_each_file(RESULT_DIR, FORMAT_FILE)
    print(DATA_DIR)
    if recognization:
        print(f"Error! Trader or Pipeline may not be Completed in these files: {recognization}")
    else:
        print(f"Success! Trader and Pipeline all Recognized")


if __name__ == "__main__":
    main()